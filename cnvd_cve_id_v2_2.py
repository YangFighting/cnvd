#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import datetime
import os
import time

import openpyxl
import xlsxwriter
from selenium.webdriver.chrome.options import Options
from selenium import webdriver
import threading
from queue import Queue

##爬取 cnvd 网站的漏洞  （多线程）##

class CnvdSpdier():
    def __init__(self):
        self.url_lines = []
        self.vulnerability_num = 0
        self.vulnerability_twebdriverotal_list = []
        self.url_queue = Queue()
        self.html_queue = Queue()
        self.w_data_queue = Queue()
        self.failure_urls_id_list =[]

        self.filetime = time.strftime("%Y-%m-%d-%H_%M_%S", time.localtime(time.time()))
        self.current_file_name = os.path.split(__file__)[-1].split(".")[0]

        self.excel_file_name = 'spider_' + self.current_file_name + '_' + self.filetime + '.xlsx'
        self.sheetname = 'cnvd_vulnerability'
        self.excel_folder_name = 'excel'
        self.path_excel_file_name = os.path.join(os.getcwd(),
                                                        self.excel_folder_name, self.excel_file_name)

        self.failure_urls_name = 'failure_urls_' + self.current_file_name + '_' + self.filetime + '.txt'
        self.failure_urls_folder_name = 'failure_urls'
        self.path_failure_urls_file_name = os.path.join(os.getcwd(),
                                            self.failure_urls_folder_name, self.failure_urls_name)

        self.info_dict_folder_name = 'info_dict'
        if not os.path.exists(self.info_dict_folder_name):
            os.mkdir(self.info_dict_folder_name)

        if not os.path.exists(self.failure_urls_folder_name):
            os.mkdir(self.failure_urls_folder_name)

        if not os.path.exists(self.excel_folder_name):
            os.mkdir(self.excel_folder_name)

        # 表头
        self.excel_head = ['漏洞名称', 'CNVD-ID', '公开日期', '危害级别', '影响产品', 'CVE ID', '漏洞描述',
                      '漏洞类型', '参考链接', '漏洞解决方案', '厂商补丁', '报送时间', '收录时间', '更新时间']

    def read_txt(self, file_name='isc_url.txt'):
        with open(file_name, "r") as f:
            self.url_lines = f.readlines()

    def vulnerability_get(self, url=None):
        """
            模拟浏览器，请求URL
        :param url:
        :return:
        """
        # 创建chrome浏览器驱动，无头模式
        chrome_options = Options()
        chrome_options.add_argument('--headless')
        is_except = 0

        try:
            driver = webdriver.Chrome(chrome_options=chrome_options)
            driver.get(url)
            time.sleep(2)
        except:
            is_except = 1

        sleep_time = 2.5
        # 访问次数过多重置浏览器
        while is_except == 1:
            try:
                driver.quit()
                driver = webdriver.Chrome(chrome_options=chrome_options)
                driver.get(url)
                time.sleep(sleep_time)
                is_except = 0
            except:
                is_except = 1
                print('-' * 20 + '%d' % sleep_time + '-' * 20 )
                sleep_time = sleep_time*2

            if sleep_time > pow(2, 16):  # 长时间不能登录
                print('-' * 20 + '长时间不能登录' + '-' * 20)
                return None

        vulnerability_html_dict = {}
        title = driver.find_element_by_css_selector('div.blkContainerSblk > h1').text
        vulnerability_html_dict[self.excel_head[0]] = title

        css_selector_tr_list = driver.find_elements_by_css_selector('table.gg_detail > tbody > tr')
        for css_selector_tr_i in css_selector_tr_list:
            vulnerability_html_kv = css_selector_tr_i.find_elements_by_css_selector('td')
            if len(vulnerability_html_kv) == 2:
                vulnerability_html_key = vulnerability_html_kv[0].text
                vulnerability_html_dict[vulnerability_html_key] = vulnerability_html_kv[1].text

        driver.quit()
        # file_name = 'hml_str_' + str(self.vulnerability_num) + '.txt'
        # self.save_hml_dict(file_name=file_name, hml_dict=vulnerability_html_dict)

        return vulnerability_html_dict

    def parse_web_page(self, vulnerability_html_dict=None):
        """
            解析 网页字典，用于写入excel
        :param vulnerability_html_dict:
        :return:
        """
        vulnerability_list = []
        info_dict ={}

        for (hml_dict_k, hml_dict_v) in vulnerability_html_dict.items():
            if hml_dict_k in self.excel_head:
                info_dict[hml_dict_k] = hml_dict_v

        if not info_dict.get('CVE ID'):     #CVE ID不存在
            info_dict['CVE ID']= ''

        # 按照 表头顺序 添加到病毒信息列表中
        for title_i in self.excel_head:
            vulnerability_list.append(info_dict[title_i])

        return vulnerability_list

    def vulnerability_excel_init(self):
        """
            漏洞信息excel 表头 初始化
        """
        # workbook = xlsxwriter.Workbook(self.excel_file_name)
        workbook = xlsxwriter.Workbook(self.path_excel_file_name)
        worksheet = workbook.add_worksheet(self.sheetname)

        for col_i in range(0, len(self.excel_head)):
            worksheet.write(0, col_i, self.excel_head[col_i])
        workbook.close()

    def vulnerability_write_excel(self, write_data_list=None):
        """
            单条漏洞信息写入excel
        """
        workbook = openpyxl.load_workbook(self.path_excel_file_name)
        worksheet = workbook[self.sheetname]
        rows_old = worksheet.max_row  # 获取表格中已存在的数据的行数
        for col_i in range(0, len(write_data_list)):
            worksheet.cell(rows_old+1, col_i+1).value = write_data_list[col_i]

        workbook.save(self.path_excel_file_name)

    def save_hml_dict(self, file_name, hml_dict = None):
        path_info_dict_file_name = os.path.join( os.getcwd(), self.info_dict_folder_name, file_name)
        with open(path_info_dict_file_name, "w") as f:
            for (hml_dict_k, hml_dict_v) in hml_dict.items():
                f.write(hml_dict_k)
                f.write('\n')
                f.write(hml_dict_v)
                f.write('\n')

    def write_failure_url(self, failure_url=None):

        with open(self.path_failure_urls_file_name, "a+") as f:
            f.writelines(failure_url)

    def get_url_list(self):
        # 读文件,获得URl
        self.read_txt()
        # self.url_lines = self.url_lines[0:6]

        self.url_lines = ['https://www.cnvd.org.cn/flaw/show/CNVD-2019-16786']
        for url_i in self.url_lines:
            self.url_queue.put(url_i)

    def request_vulnerability_url(self):
        """
             模拟浏览器, 请求URL
        """
        while True:
            url = self.url_queue.get()
            try:
                html_dict = self.vulnerability_get(url=url)
                self.html_queue.put(html_dict)
            except Exception as err:
                print('Exception: ', err)
                self.write_failure_url(failure_url=url)

            self.url_queue.task_done()
            print('爬取： %s' % str(url).replace('\n', ''))

    def get_content_list(self):
        """
            解析数据
        """
        while True:
            html_dict = self.html_queue.get()
            try:
                vulnerability_list = self.parse_web_page(vulnerability_html_dict=html_dict)
            except Exception as err:
                print('Exception: ', err)
                vulnerability_list = None
                self.write_failure_url(failure_url=self.url_lines[self.vulnerability_num])
                print('爬取失败： 第 %d 个 ' % (self.vulnerability_num + 1))

            if vulnerability_list:
                self.w_data_queue.put(vulnerability_list)
                print('已爬取：第 %d 个 ' % (self.vulnerability_num + 1))
            self.html_queue.task_done()
            self.vulnerability_num += 1

    def save_content_list(self):
        """
            保存数据
        """
        while True:
            content_list = self.w_data_queue.get()
            self.vulnerability_write_excel(write_data_list=content_list)
            self.w_data_queue.task_done()

    def run(self):
        """
            实现 主要逻辑
        """
        self.vulnerability_excel_init()

        # 计算 爬虫时间
        begin = datetime.datetime.now()

        thread_list = []
        t_url = threading.Thread(target=self.get_url_list)
        thread_list.append(t_url)

        for i in range(8):
            t_request = threading.Thread(target=self.request_vulnerability_url)
            thread_list.append(t_request)

        for i in range(5):
            t_html = threading.Thread(target=self.get_content_list)
            thread_list.append(t_html)

        t_save = threading.Thread(target=self.save_content_list)
        thread_list.append(t_save)

        for t in thread_list:
            t.setDaemon(True)  # 把子线程设置为守护线程，该线程不重要, 主线程结束，子线程结束
            t.start()

        for q in [self.url_queue, self.html_queue, self.w_data_queue]:
            q.join()  # 让主线程等待阻塞，等待队列的任务完成之后再完成

        # 爬行结束
        end = datetime.datetime.now()
        total_time = end - begin
        print('漏洞信息爬取结束')
        print('爬行时间： ', total_time)


if __name__ == "__main__":
    cnvd_spdier = CnvdSpdier()
    cnvd_spdier.run()
